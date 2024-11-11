import openpyxl as op
wb=op.load_workbook("person_test.xlsx")
ws=wb.active
ws["C4"].value= "=Sum(C2:C3)" #C4: 단가의 합
ws["D4"].value= "=Sum(D2:D3)" # D4:인원의합
ws["C5"].value= "=AVERAGE(C2:C3)" # C5:단가 의평균
ws["D5"].value= "=AVERAGE(D2:D3)" #  D5:인원의 평균
wb.save("person_test.xlsx") # 저장

