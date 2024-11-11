import openpyxl as op
def customer_number():
    ws=wb.active
    list_customer_numbe=list(map(int,input().split()))
    for i in range(len(list_customer_numbe)):
        ws.cell(row=i+2,column=1).value=list_customer_numbe[i]
def customer_name():
    ws=wb.active
    list_customer_name=list(input().split())
    for i in range(len(list_customer_name)):
        ws.cell(row=i+2,column=2).value=list_customer_name[i]
def Number_of_payments():
    ws=wb.active
    list_Number_of_payment=list(map(int,input().split()))
    for i in range(len(list_Number_of_payment)):
        ws.cell(row=i+2,column=3).value=list_Number_of_payment[i]    
def Amount_used():
    ws=wb.active
    list_Amount_used=list(map(int,input().split()))
    for i in range(len(list_Amount_used)):
        ws.cell(row=i+2,column=4).value=list_Amount_used[i]
def Customer_rating():
    ws=wb.active
    list_Customer_rating=list(input().split())
    for i in range(len(list_Customer_rating)):
        ws.cell(row=i+2,column=5).value=list_Customer_rating[i] 
def Number_of_overdue_payments():
    ws=wb.active
    list_Number_of_overdue_payments=list(map(int,input().split()))
    for i in range(len(list_Number_of_overdue_payments)):
        ws.cell(row=i+2,column=6).value=list_Number_of_overdue_payments[i]          
def main():
    global wb
    wb=op.Workbook()
    ws=wb.active
    lisr_customer=list(input().split())
    for i in range(len(lisr_customer)):
        ws.cell(row=1,column=i+1).value=lisr_customer[i]
    customer_number()
    customer_name()
    Number_of_payments()
    Amount_used()
    Customer_rating()
    Number_of_overdue_payments()
    wb.save("card_csv.xlsx")
main()