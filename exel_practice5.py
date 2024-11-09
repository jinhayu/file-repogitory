"""""
실습문제 (5)
조건: "openpyxl_test.xlsx" 파일의 활성화된 Sheet에서 A1부터 C3까지의 모든 셀의 값을 출력하시오.
"""
import openpyxl as op

def print1():
    ws=wb.active
    rag=ws["A1:C3"]
    for rng_data in rag:
        for cell_data in rng_data:
            print(cell_data.value)

def main():
    global wb
    wb=op.load_workbook("openpyxl_test.xlsx")
    ws=wb.active
    for i in range(1,4):
        for j in range(1,4):
            ws.cell(row=i,column=j).value=i

    print1()
    wb.save("openpyxl_test.xlsx")
main()