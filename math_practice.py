import openpyxl as op
def Name():
    ws=wb.active
    list_customer_numbe=list(input().split())
    for i in range(len(list_customer_numbe)):
        ws.cell(row=i+2,column=1).value=list_customer_numbe[i]
def gucku():
    ws=wb.active
    list_customer_name=list(map(float,input().split()))
    for i in range(len(list_customer_name)):
        ws.cell(row=i+2,column=2).value=list_customer_name[i]
def math():
    ws=wb.active
    list_Number_of_payment=list(map(float,input().split()))
    for i in range(len(list_Number_of_payment)):
        ws.cell(row=i+2,column=3).value=list_Number_of_payment[i]    
def english():
    ws=wb.active
    list_Amount_used=list(map(float,input().split()))
    for i in range(len(list_Amount_used)):
        ws.cell(row=i+2,column=4).value=list_Amount_used[i]
      
def main():
    global wb
    wb=op.Workbook()
    ws=wb.active
    lisr_customer=list(input().split())
    for i in range(len(lisr_customer)):
        ws.cell(row=1,column=i+1).value=lisr_customer[i]
    Name()
    gucku()
    math()
    english()
    wb.save("math practice.xlsx")
main()