import openpyxl as op  

def Name():  # 이름 입력 
    ws = wb.active 
    list_customer_numbe = list(input().split())  # 사용자 입력을 받아 리스트로 저장
    for i in range(len(list_customer_numbe)):
        ws.cell(row=i+2, column=1).value = list_customer_numbe[i]  # A열(1열)에 이름 저장

def date():  # 받은 날짜 입력 
    ws = wb.active
    list_customer_numbe = list(input().split())  # 사용자 입력을 받아 리스트로 저장
    for i in range(len(list_customer_numbe)):
        ws.cell(row=i+2, column=2).value = list_customer_numbe[i]  # B열(2열)에 받은 날짜 저장

def check():  # 수령 여부 입력 
    ws = wb.active
    list_customer_numbe = list(input().split())  # 사용자 입력을 받아 리스트로 저장
    for i in range(len(list_customer_numbe)):
        ws.cell(row=i+2, column=3).value = list_customer_numbe[i]  # C열(3열)에 수령 여부 저장

def main():  # 메인 함수
    global wb 
    wb = op.Workbook()  
    ws = wb.active  
    
    lisr_customer = list(input().split())  # 첫 번째 줄에 들어갈 헤더 입력 (예: "이름 날짜 수령여부")
    for i in range(len(lisr_customer)):
        ws.cell(row=1, column=i+1).value = lisr_customer[i]  # 첫 번째 행에 헤더 추가
    
    Name()  #  이름 입력  실행
    date()  # 받은 날짜 입력  실행
    check()  # 수령 여부 입력  실행
    
    wb.save("money_check.xlsx")  # 입력된 데이터를 'money_check.xlsx' 파일로 저장

main()  # 메인 함수 실행