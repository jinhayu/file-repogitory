import openpyxl as op
def date(): # 날짜 열 내용 채우기 
    ws=wb.active
    list_date=list(input().split())
    for i in range(len(list_date)):
        ws.cell(row=i+2,column=1).value=list_date[i]
def peoduce(): # 항목 열 내용 채우기
    ws=wb.active
    list_produce=list(input().split())
    for i in range(len(list_produce)):
        ws.cell(row=i+2,column=2).value=list_produce[i]
def Unit_price(): # 단가 열채우기
    ws=wb.active
    list_unit_price=list(map(int,input().split()))
    for i in range(len(list_unit_price)):
        ws.cell(row=i+2,column=3).value=list_unit_price[i]
def people(): # 인원 열 채우기
     ws=wb.active
     list_peple=list(map(int,input().split()))
     for i in range(len(list_peple)):
        ws.cell(row=i+2,column=4).value=list_peple[i]
def main():
    global wb # 전역변수
    wb=op.Workbook()
    ws=wb.active
    list_1=list(input().split()) # 날짜 항목 단가 인원 총가격(단가*인원)
    for i in range(len(list_1)):
        ws.cell(row=1,column=i+1).value=list_1[i]
    date() # 날짜
    peoduce() # 항목
    Unit_price() # 단가
    people() # 인원
    ws["E2"].value= "=PRODUCT(C2,D2)" #c2*D2
    ws["E3"].value= "=PRODUCT(C3,D3)" # C3*D3
    wb.save("person_test.xlsx")
main()